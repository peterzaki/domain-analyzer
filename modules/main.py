import concurrent.futures, openpyxl, os, sys
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from datetime import datetime
from modules import process_domain as p
from modules import argParser

def main():
    try:
        # Parse command-line arguments
        args = argParser.get_args()
        verbose = args.quiet

        # Read domains from input file
        file_path = args.input
        with open(file_path, 'r') as file:
            domains = file.read().splitlines()

        results = []
        age_results = []
        dga_results = []
        dns_results = []
        ttl_results = []
        sus_results = []
        invalid = []
        
        print("Analyzing ..")
        with concurrent.futures.ThreadPoolExecutor() as executor:
            results = [executor.submit(p.process_domain, domain, verbose) for domain in domains]

            for future in concurrent.futures.as_completed(results):
                domain, age, dga_status, ip_count, ip_ttl = future.result()
                suspicion_level = ""
                concern = ""

                if (age == None) and (dga_status == None) and (ip_count == None) and (ip_ttl == None):
                    invalid.append((domain,))

                if age is not None and age <= args.days:
                    age_results.append((domain, age))
                    suspicion_level = "Low"
                    concern = "Age"

                if dga_status:
                    dga_results.append((domain,))
                    if suspicion_level == "Low":
                       suspicion_level = "Medium"
                       concern += ", DGA"
                    else:
                        suspicion_level = "Low"
                        concern = "DGA"

                if ip_count is not None and ip_count >= args.ips:
                    if ip_ttl is not None and ip_ttl <= args.ttl:
                        dns_results.append((domain, ip_count, ip_ttl))
                        if suspicion_level == "Low":
                            suspicion_level = "High"
                            concern += ", IPs & TTL"
                        elif suspicion_level == "Medium":
                            suspicion_level = "Very High"
                            concern += ", IPs & TTL"
                        else:
                            suspicion_level = "Medium"
                            concern = "IPs & TTL"
                    elif suspicion_level == "Low":
                        suspicion_level = "Medium"
                        concern += ", IPs"
                    elif suspicion_level == "Medium":
                        suspicion_level = "High"
                        concern += ", IPs"
                    else:
                        suspicion_level = "Low"
                        concern = "IPs"
    
                sus_results.append((domain, age, dga_status, ip_count, ip_ttl, suspicion_level, concern))

        # Save results to an Excel Workbook
        wb = Workbook()

        header_fill = PatternFill(start_color="245062", end_color="245062", fill_type="solid")
        header_font_color = Font(color="FFFFFF", bold=True)
        cell_alignment = Alignment(horizontal="center", vertical="center")

        # Styles dictionary for header cells
        header_styles = {'fill': header_fill,
                         'font': header_font_color,
                         'alignment': cell_alignment}

        border_style = Border(left=Side(border_style="thin"),
                              right=Side(border_style="thin"),
                              top=Side(border_style="thin"),
                              bottom=Side(border_style="thin"))

        # Newly Created Results Sheet
        if age_results:
            ws = wb.create_sheet(title="Newly Created")
            ws.append(["Domain", "Age (Days)"])
            for cell in ws['1']:
                for style_property, style_value in header_styles.items():
                    setattr(cell, style_property, style_value)
            for result in age_results:
                ws.append(result)

        # DGA Results Sheet
        if dga_results:
            ws = wb.create_sheet(title="Potential DGA")
            ws.append(["Domain"])
            for cell in ws['1']:
                for style_property, style_value in header_styles.items():
                    setattr(cell, style_property, style_value)
            for result in dga_results:
                ws.append(result)

        # Fast-Flux Results Sheet
        if dns_results:
            ws = wb.create_sheet(title="Potential Fast-Flux")
            ws.append(["Domain", "Num of IPs", "TTL"])
            for cell in ws['1']:
                for style_property, style_value in header_styles.items():
                    setattr(cell, style_property, style_value)
            for result in dns_results:
                ws.append(result)

        # Overview Results Sheet
        if sus_results:
            ws = wb.create_sheet(title="Overview")
            ws.append(["Domain", "Age (Days)", "DGA", "Num of IPs", "TTL", "Suspicion", "Concern"])
            for cell in ws['1']:
                for style_property, style_value in header_styles.items():
                    setattr(cell, style_property, style_value)
            for result in sus_results:
                ws.append(result)

        # Invalid Results Sheet
        if invalid:
            ws = wb.create_sheet(title="Invalid")
            ws.append(["Value"])
            for cell in ws['1']:
                for style_property, style_value in header_styles.items():
                    setattr(cell, style_property, style_value)
            for result in invalid:
                ws.append(result)

        for sheet in wb.sheetnames:
            print("Please wait for the final processing ..")
            try:
                if sheet == "Sheet":
                    del wb[sheet]
                else:
                    ws = wb[sheet]
                    max_cell_size = 0
                    for row in ws.iter_rows():
                        for cell in row:
                            cell.border = border_style
                            if (cell.column_letter == 'A'):
                                if max_cell_size < len(str(cell.value)):
                                    ws.column_dimensions[cell.column_letter].width = len(str(cell.value)) + 1
                                    max_cell_size = len(str(cell.value))
                            else:
                                cell.alignment = cell_alignment
                                ws.column_dimensions[cell.column_letter].width = 10
            except:
                print("Error!!")

        # Check if there is at least one non-empty sheet
        if not wb.sheetnames:
            print("No results found.")
        else:
            # Generate a unique filename with the current date and time
            output_filename = datetime.now().strftime("DA-Results_%Y-%m-%d_%H-%M.xlsx")
            output_file = os.path.join(os.getcwd(), output_filename)
    
            # Save the results to the generated filename
            wb.save(output_file)
            print(f'Results saved to "{output_filename}".')

    except KeyboardInterrupt:
        print("Script interrupted by the user. Exiting...")
        sys.exit(0)
